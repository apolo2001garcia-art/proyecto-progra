
from openpyxl import Workbook, load_workbook
import os
from modelo.rzr import RZR


class ExcelManager:

    def __init__(self, archivo="rzr_datos.xlsx"):

        self.archivo = archivo

        if not os.path.exists(self.archivo):

            wb = Workbook()

           

            hoja = wb.active
            hoja.title = "RZR"

            hoja.append([
                "Modelo",
                "Precio",
                "Tipo",
                "Cantidad",
                "Activo"
            ])


            hoja_clientes = wb.create_sheet(
                "Clientes"
            )

            hoja_clientes.append([
                "Nombre",
                "Telefono",
                "Correo"
            ])

       
            hoja_operaciones = wb.create_sheet(
                "Operaciones"
            )

            hoja_operaciones.append([
                "Cliente",
                "Vehiculo",
                "Operacion",
                "Precio"
            ])

            wb.save(self.archivo)


    def actualizar_rzr(self, rzr_actualizado):

        wb = load_workbook(self.archivo)

        hoja = wb["RZR"]

        for fila in hoja.iter_rows(min_row=2):

            modelo_excel = fila[0].value

            if modelo_excel == rzr_actualizado.get_modelo():

                fila[1].value = rzr_actualizado.get_precio()
                fila[2].value = rzr_actualizado.get_tipo()
                fila[3].value = rzr_actualizado.get_cantidad()
                fila[4].value = rzr_actualizado.get_activo()

                break

        wb.save(self.archivo)

   

    def guardar_rzr(self, rzr):

        wb = load_workbook(self.archivo)

        hoja = wb["RZR"]

        hoja.append([
            rzr.get_modelo(),
            rzr.get_precio(),
            rzr.get_tipo(),
            rzr.get_cantidad(),
            rzr.get_activo()
        ])

        wb.save(self.archivo)

  

    def cargar_rzrs(self):

        wb = load_workbook(self.archivo)

        hoja = wb["RZR"]

        rzrs = []

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            modelo, precio, tipo, cantidad, activo = fila

            rzr = RZR(
                modelo,
                precio,
                tipo,
                cantidad,
                activo
            )

            rzrs.append(rzr)

        return rzrs

   

    def guardar_cliente(self, cliente):

        wb = load_workbook(self.archivo)

        hoja = wb["Clientes"]

        hoja.append([
            cliente.get_nombre(),
            cliente.get_telefono(),
            cliente.get_correo()
        ])

        wb.save(self.archivo)

    

    def cargar_clientes(self):

        from modelo.cliente import Cliente

        wb = load_workbook(self.archivo)

        hoja = wb["Clientes"]

        clientes = []

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            nombre, telefono, correo = fila

            cliente = Cliente(
                nombre,
                telefono,
                correo
            )

            clientes.append(cliente)

        return clientes

    

    def guardar_operacion(
        self,
        cliente,
        vehiculo,
        operacion,
        precio
    ):

        wb = load_workbook(self.archivo)

        hoja = wb["Operaciones"]

        hoja.append([
            cliente,
            vehiculo,
            operacion,
            precio
        ])

        wb.save(self.archivo)