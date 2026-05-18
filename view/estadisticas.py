import customtkinter as ctk
from openpyxl import load_workbook


class VentanaEstadisticas(ctk.CTkToplevel):

    def __init__(self, parent):
        super().__init__(parent)

        self.focus()
        self.lift()
        self.attributes("-topmost", True)
        self.after(100, lambda: self.attributes("-topmost", False))

        self.title("Estadísticas")
        self.geometry("700x550")

        

        ctk.CTkLabel(
            self,
            text="Estadísticas del Sistema",
            font=("Arial", 28, "bold")
        ).pack(pady=20)

        

        total_ingresos = self.calcular_ingresos()
        total_operaciones = self.contar_operaciones()
        total_ventas = self.contar_tipo("Venta")
        total_rentas = self.contar_tipo("Renta")

        vehiculo_top = self.vehiculo_mas_usado()

        inventario = self.calcular_inventario()

        

        frame = ctk.CTkFrame(self)

        frame.pack(
            padx=20,
            pady=20,
            fill="both",
            expand=True
        )

       

        ingresos_frame = ctk.CTkFrame(frame)

        ingresos_frame.pack(
            pady=10,
            padx=10,
            fill="x"
        )

        ctk.CTkLabel(
            ingresos_frame,
            text="Ingresos Totales",
            font=("Arial", 22, "bold")
        ).pack(pady=10)

        ctk.CTkLabel(
            ingresos_frame,
            text=f"${total_ingresos:,.2f}",
            font=("Arial", 30)
        ).pack(pady=10)

        

        operaciones_frame = ctk.CTkFrame(frame)

        operaciones_frame.pack(
            pady=10,
            padx=10,
            fill="x"
        )

        ctk.CTkLabel(
            operaciones_frame,
            text=f"Operaciones registradas: {total_operaciones}",
            font=("Arial", 18)
        ).pack(pady=5)

        ctk.CTkLabel(
            operaciones_frame,
            text=f"Ventas realizadas: {total_ventas}",
            font=("Arial", 18)
        ).pack(pady=5)

        ctk.CTkLabel(
            operaciones_frame,
            text=f"Rentas realizadas: {total_rentas}",
            font=("Arial", 18)
        ).pack(pady=5)

        ctk.CTkLabel(
            operaciones_frame,
            text=f"Inventario disponible: {inventario}",
            font=("Arial", 18)
        ).pack(pady=5)

        

        top_frame = ctk.CTkFrame(frame)

        top_frame.pack(
            pady=10,
            padx=10,
            fill="x"
        )

        ctk.CTkLabel(
            top_frame,
            text="Vehículo más solicitado",
            font=("Arial", 20, "bold")
        ).pack(pady=10)

        ctk.CTkLabel(
            top_frame,
            text=vehiculo_top,
            font=("Arial", 22)
        ).pack(pady=10)

   

    def calcular_ingresos(self):

        wb = load_workbook("rzr_datos.xlsx")

        hoja = wb["Operaciones"]

        total = 0

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            precio = fila[3]

            if precio is not None:
                total += precio

        return total

    

    def contar_operaciones(self):

        wb = load_workbook("rzr_datos.xlsx")

        hoja = wb["Operaciones"]

        contador = 0

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            if fila[0] is not None:
                contador += 1

        return contador

   

    def contar_tipo(self, tipo_busqueda):

        wb = load_workbook("rzr_datos.xlsx")

        hoja = wb["Operaciones"]

        contador = 0

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            tipo = fila[2]

            if tipo == tipo_busqueda:
                contador += 1

        return contador

    

    def vehiculo_mas_usado(self):

        wb = load_workbook("rzr_datos.xlsx")

        hoja = wb["Operaciones"]

        conteo = {}

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            vehiculo = fila[1]

            if vehiculo is not None:

                if vehiculo in conteo:
                    conteo[vehiculo] += 1

                else:
                    conteo[vehiculo] = 1

        if not conteo:
            return "Sin operaciones"

        return max(
            conteo,
            key=conteo.get
        )

    

    def calcular_inventario(self):

        wb = load_workbook("rzr_datos.xlsx")

        hoja = wb["RZR"]

        total = 0

        for fila in hoja.iter_rows(
            min_row=2,
            values_only=True
        ):

            cantidad = fila[3]

            if cantidad is not None:
                total += cantidad

        return total